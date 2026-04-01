"""Format an InvestmentMemo into a readable report.

Outputs to terminal (via rich) and/or markdown file, and fills the
Excel underwriting template with property-specific data.
"""

from __future__ import annotations

import subprocess
from pathlib import Path

from models import InvestmentMemo

# Locate key paths relative to this file (src/ → project root)
_SRC_DIR      = Path(__file__).parent
_PROJECT_ROOT = _SRC_DIR.parent
_TEMPLATE_XLS = _PROJECT_ROOT / "templates" / "STR_Underwriting_Template.xlsx"
_MEMOS_DIR    = _PROJECT_ROOT / "memos"
_RECALC       = _PROJECT_ROOT.parent / ".skills" / "skills" / "xlsx" / "scripts" / "recalc.py"


def to_markdown(memo: InvestmentMemo) -> str:
    """Render a full investment memo as markdown."""
    p = memo.property
    m = memo.market
    r = memo.revenue
    e = memo.expenses
    f = memo.financing
    ret = memo.returns
    t = memo.tax

    lines: list[str] = []

    # ── Header ────────────────────────────────────────────────────────────
    lines.append(f"# STR Investment Memo — {p.address}")
    lines.append(f"**Analysis Date:** {memo.analysis_date}  ")
    lines.append(f"**Data Confidence:** {memo.data_confidence.upper()}  ")
    lines.append(f"**Recommendation:** **{memo.recommendation.value}**\n")

    # ── Property Summary ──────────────────────────────────────────────────
    lines.append("## 1. Property Summary\n")
    lines.append(f"| Field | Value |")
    lines.append(f"|-------|-------|")
    lines.append(f"| Address | {p.address} |")
    lines.append(f"| Market | {p.market_name or p.city + ', ' + p.state} |")
    lines.append(f"| Price | ${p.price:,.0f} |")
    lines.append(f"| Beds / Baths | {p.bedrooms} / {p.bathrooms} |")
    lines.append(f"| Sqft | {p.sqft:,} |")
    if p.year_built:
        lines.append(f"| Year Built | {p.year_built} |")
    if p.listing_url:
        lines.append(f"| Listing | [Zillow]({p.listing_url}) |")
    lines.append("")

    # ── Revenue Estimate ──────────────────────────────────────────────────
    lines.append("## 2. Revenue Estimate\n")
    lines.append(f"| Metric | Value |")
    lines.append(f"|--------|-------|")
    lines.append(f"| ADR | ${r.adr:,.0f} |")
    lines.append(f"| Occupancy | {r.occupancy_rate:.1%} |")
    lines.append(f"| Nights Booked | {r.total_nights_booked:.0f} |")
    lines.append(f"| Turnovers | {r.total_turnovers:.0f} |")
    lines.append(f"| **Gross Revenue** | **${r.gross_revenue:,.0f}** |")
    lines.append(f"\n*Data source: {m.data_source} | Confidence: {m.confidence}*\n")

    # ── Expense Breakdown ─────────────────────────────────────────────────
    lines.append("## 3. Expense Breakdown\n")
    lines.append(f"| Expense | Annual | % of Revenue |")
    lines.append(f"|---------|--------|-------------|")

    expense_items = [
        ("Platform Fees", e.platform_fees),
        ("Property Management", e.property_management),
        ("Cleaning", e.cleaning),
        ("Supplies", e.supplies),
        ("Utilities", e.utilities),
        ("Insurance", e.insurance),
        ("Property Tax", e.property_tax),
        ("HOA", e.hoa),
        ("Maintenance Reserve", e.maintenance_reserve),
        ("Lawn / Snow", e.lawn_snow),
        ("Pest Control", e.pest_control),
        ("Hot Tub / Pool", e.hot_tub_pool),
        ("Permits & Licenses", e.permits_licenses),
        ("Accounting", e.accounting),
        ("Software", e.software),
    ]
    for name, val in expense_items:
        pct = val / r.gross_revenue if r.gross_revenue else 0
        lines.append(f"| {name} | ${val:,.0f} | {pct:.1%} |")

    lines.append(f"| **Total Expenses** | **${e.total_expenses:,.0f}** | **{e.expense_ratio:.1%}** |")
    lines.append("")

    # ── Financing ─────────────────────────────────────────────────────────
    lines.append("## 4. Financing\n")
    lines.append(f"| Field | Value |")
    lines.append(f"|-------|-------|")
    lines.append(f"| Loan Type | {f.loan_type} |")
    lines.append(f"| Down Payment | ${f.down_payment:,.0f} ({f.down_payment/f.purchase_price:.0%}) |")
    lines.append(f"| Loan Amount | ${f.loan_amount:,.0f} |")
    lines.append(f"| Interest Rate | {f.interest_rate:.2%} |")
    lines.append(f"| Monthly P&I | ${f.monthly_payment:,.0f} |")
    lines.append(f"| Annual Debt Service | ${f.annual_debt_service:,.0f} |")
    lines.append(f"| Closing Costs | ${f.closing_costs:,.0f} |")
    lines.append(f"| **Total Cash to Close** | **${f.total_cash_required:,.0f}** |")
    lines.append("")

    # ── Return Metrics ────────────────────────────────────────────────────
    lines.append("## 5. Return Metrics\n")
    lines.append(f"| Metric | Value |")
    lines.append(f"|--------|-------|")
    lines.append(f"| NOI | ${ret.noi:,.0f} |")
    lines.append(f"| Cash Flow (pre-tax) | ${ret.cash_flow_before_tax:,.0f} |")
    lines.append(f"| **Cash-on-Cash Return** | **{ret.cash_on_cash_return:.1%}** |")
    lines.append(f"| Cap Rate | {ret.cap_rate:.1%} |")
    lines.append(f"| Gross Yield | {ret.gross_yield:.1%} |")
    lines.append(f"| DSCR | {ret.dscr:.2f} |")
    if ret.year3_roe is not None:
        lines.append(f"| Year 3 ROE | {ret.year3_roe:.1%} |")
    if ret.year5_roe is not None:
        lines.append(f"| Year 5 ROE | {ret.year5_roe:.1%} |")
    lines.append("")

    # ── Tax Impact ────────────────────────────────────────────────────────
    lines.append("## 6. Tax Impact (Cost Segregation)\n")
    lines.append(f"| Component | Basis | Dep. Schedule |")
    lines.append(f"|-----------|-------|---------------|")
    lines.append(f"| Land (non-depreciable) | ${t.land_value:,.0f} | — |")
    lines.append(f"| Personal Property | ${t.personal_property_basis:,.0f} | 5/7-year |")
    lines.append(f"| Land Improvements | ${t.land_improvement_basis:,.0f} | 15-year |")
    lines.append(f"| Building | ${t.building_basis:,.0f} | 27.5-year |")
    lines.append("")
    lines.append(f"| Tax Metric | Federal | California |")
    lines.append(f"|------------|---------|------------|")
    lines.append(f"| Year 1 Depreciation | ${t.year1_total_depreciation:,.0f} | ${t.year1_ca_depreciation:,.0f} |")
    lines.append(f"| Tax Savings | ${t.federal_tax_savings:,.0f} | ${t.ca_tax_savings:,.0f} |")
    lines.append(f"| **Total Year 1 Tax Savings** | **${t.total_tax_savings:,.0f}** | |")
    lines.append(f"| Effective Year 1 Cost | ${t.effective_year1_cost:,.0f} | |")
    lines.append(f"\n*Note: California does not conform to federal bonus depreciation.*\n")

    # ── Risk Factors ──────────────────────────────────────────────────────
    lines.append("## 7. Risk Factors\n")
    for risk in memo.risks:
        icon = {"high": "🔴", "medium": "🟡", "low": "🟢"}[risk.severity]
        lines.append(f"- {icon} **{risk.category.title()}**: {risk.description}")
    lines.append("")

    # ── Recommendation ────────────────────────────────────────────────────
    lines.append("## 8. Recommendation\n")
    emoji = {
        "STRONG BUY": "🟢🟢",
        "BUY": "🟢",
        "HOLD": "🟡",
        "PASS": "🔴",
    }[memo.recommendation.value]
    lines.append(f"### {emoji} {memo.recommendation.value}\n")
    lines.append(memo.recommendation_rationale.replace(" | ", "\n\n"))
    lines.append("")

    lines.append("---")
    lines.append("*Generated by STR Investment Analyzer. Not financial advice.*")

    return "\n".join(lines)


def save_markdown(memo: InvestmentMemo, path: str) -> None:
    """Write the memo to a markdown file."""
    md = to_markdown(memo)
    with open(path, "w") as f:
        f.write(md)
    print(f"Memo saved to {path}")


# ── Excel spreadsheet output ──────────────────────────────────────────────

def to_xlsx(
    memo: InvestmentMemo,
    template_path: Path | str | None = None,
    output_path: Path | str | None = None,
) -> Path:
    """Fill the Excel underwriting template with data from the memo and save.

    Only the yellow (user-input) cells are written; all formula cells are
    left intact so they recalculate automatically after save.

    Input cells populated from the memo
    ─────────────────────────────────────
    Sheet: underwriting
      D2        Address
      D6        Bedrooms
      D7        Bathrooms
      D8        Square footage
      D9        Market name
      D12       Purchase price
      D31/E31/F31  ADR  (Low −15% / Mid = actual / High +12%)
      D32/E32/F32  Occupancy  (Low −10% / Mid = actual / High +7%, capped 85%)
      D36       Data source & confidence
      D45       Annual property tax (from listing)
    """
    from openpyxl import load_workbook

    template_path = Path(template_path or _TEMPLATE_XLS)
    if output_path is None:
        slug = (
            memo.property.address
            .replace(", ", "-")
            .replace(" ", "-")
            .replace("/", "-")
            .replace(",", "")[:60]
        )
        output_path = _MEMOS_DIR / f"{slug}.xlsx"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path)
    ws = wb["underwriting"]

    p   = memo.property
    rev = memo.revenue
    exp = memo.expenses
    mkt = memo.market

    # ── Property Info ────────────────────────────────────────────────────
    ws["D2"] = p.address
    ws["D6"] = p.bedrooms
    ws["D7"] = p.bathrooms
    ws["D8"] = p.sqft
    ws["D9"] = p.market_name or f"{p.city}, {p.state}"

    # ── Purchase Price ───────────────────────────────────────────────────
    ws["D12"] = p.price

    # ── Revenue scenarios ────────────────────────────────────────────────
    # Mid = actual underwritten values; Low/High are ±15% / ±12% ADR,
    # ±10% / ±7% occupancy (occupancy capped at 0.85).
    adr_mid = rev.adr
    occ_mid = rev.occupancy_rate
    ws["D31"] = round(adr_mid * 0.85)
    ws["E31"] = round(adr_mid)
    ws["F31"] = round(adr_mid * 1.12)
    ws["D32"] = round(max(occ_mid * 0.90, 0.40), 3)
    ws["E32"] = round(occ_mid, 3)
    ws["F32"] = round(min(occ_mid * 1.07, 0.85), 3)

    # ── Data source ──────────────────────────────────────────────────────
    ws["D36"] = f"{mkt.data_source.upper()}  |  confidence: {mkt.confidence}"

    # ── Property Tax (annual, from Zillow listing) ───────────────────────
    if exp.property_tax and exp.property_tax > 0:
        ws["D45"] = exp.property_tax

    wb.save(output_path)

    # ── Recalculate formulas via LibreOffice ─────────────────────────────
    if _RECALC.exists():
        result = subprocess.run(
            ["python3", str(_RECALC), str(output_path), "90"],
            capture_output=True,
            text=True,
        )
        if result.stdout:
            import json
            try:
                info = json.loads(result.stdout)
                if info.get("total_errors", 0) > 0:
                    print(f"⚠️  Spreadsheet formula errors: {info['error_summary']}")
            except json.JSONDecodeError:
                pass

    print(f"Spreadsheet saved to {output_path}")
    return output_path


def save_xlsx(
    memo: InvestmentMemo,
    template_path: Path | str | None = None,
    output_path: Path | str | None = None,
) -> Path:
    """Convenience wrapper — fills the template and saves. Returns output path."""
    return to_xlsx(memo, template_path=template_path, output_path=output_path)
